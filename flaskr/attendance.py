import os
import functools

from flask import (
    Blueprint, flash, g, redirect, render_template, request,jsonify, session, url_for,send_file
)
import json
from werkzeug.security import check_password_hash, generate_password_hash

from flaskr.db import get_db
from datetime import datetime
from flaskr.student import get_batch_list
from openpyxl import Workbook
import smtplib
from email.mime.multipart import MIMEMultipart
from email.mime.text import MIMEText
from email.mime.base import MIMEBase
from email import encoders

bp = Blueprint('attendance', __name__, url_prefix='/attendance')

PROJECT_ROOT = os.path.abspath(os.path.dirname(__file__))

EMAIL = 'er.garimasubedi@gmail.com'
PASSWORD = 'uppv fytq mcqr ztxl'


@bp.route('/list', methods=['GET'])
def index():
    db = get_db()
    attendance = db.execute(
        '''select a.id,
            b.batch_year,
            c.course_name,
            s.subject_name,
            a.attendance_date,
            t.first_name|| ' ' ||t.last_name created_by, 
            count(1) total_student,
            COUNT(CASE WHEN ast.status = 1 THEN st.id END) AS present_students,
            COUNT(CASE WHEN ast.status = 0 THEN st.id END) AS absent_students,
            a.type
            from attendance a 
            JOIN attendance_student ast ON a.id = ast.attendance_id
            JOIN batch b ON b.id = a.batch_id
            JOIN course c ON c.id = a.course_id
            JOIN subject s ON s.id = a.subject_id
            JOIN student st ON c.id = st.course_id AND ast.student_id = st.id
            JOIN teacher t ON t.user_id = a.user_id
            GROUP BY a.id,b.batch_year,c.course_name,s.subject_name,a.attendance_date'''
    ).fetchall()
    return render_template('attendance/attendancelist.html', attendance=attendance)


@bp.route('/<int:id>/list', methods=['GET'])
def student_list(id):
    db = get_db()
    students  = db.execute(
        '''select st.id,st.first_name || " "|| st.last_name student_name,ast.status
            from attendance a 
            JOIN attendance_student ast ON a.id = ast.attendance_id
            JOIN batch b ON b.id = a.batch_id
            JOIN course c ON c.id = a.course_id
            JOIN subject s ON s.id = a.subject_id
            JOIN student st ON c.id = st.course_id AND ast.student_id = st.id
            JOIN teacher t ON t.user_id = a.user_id 
            WHERE a.id = ?''',
            (id,)
    ).fetchall()
    return render_template('attendance/studentlist.html', students=students,attendance_id=id)


def get_course_list():
    post = get_db().execute(
        'SELECT c.id,c.course_name'
        ' FROM course c'
        ' INNER JOIN course_subject cs ON c.id = cs.course_id'
        ' INNER JOIN subject s ON s.id = cs.subject_id'
        ' INNER JOIN teacher t ON t.id = s.teacher_id'
        ' WHERE t.user_id = ?',
        (g.user['id'],)
    ).fetchall()
    return post


@bp.route('/get_subject_ddl', methods=['GET'])
def get_subject_list_ddl():
    course_id = request.args.get('course_id', default='default_value', type=str)
    subjects = get_db().execute(
        'SELECT s.id, s.subject_name'
        ' FROM course c'
        ' INNER JOIN course_subject cs ON c.id = cs.course_id'
        ' INNER JOIN subject s ON s.id = cs.subject_id'
        ' INNER JOIN teacher t ON t.id = s.teacher_id'
        ' WHERE c.id = ? and t.user_id = ?',
        (course_id,g.user['id'],)
    ).fetchall()
    subject_list = [{'id': subject['id'],'subject_name': subject['subject_name']} for subject in subjects]
    return jsonify(subject_list)


@bp.route('/takeattendance', methods=('GET', 'POST'))
def takeattendance():
    batch_list = get_batch_list()
    course_list = get_course_list()
    today_date = datetime.today().date().strftime('%Y-%m-%d')
    return render_template('attendance/selectattendance.html',today_date=today_date,batch_list = batch_list,course_list=course_list)


@bp.route('/checkattendance', methods=['GET'])
def check_attendance():
    batch_id = request.args.get('batch_id', default='default_value', type=str)
    course_id = request.args.get('course_id', default='default_value', type=str)
    subject_id = request.args.get('subject_id', default='default_value', type=str)
    attendance_date = request.args.get('attendance_date', default='default_value', type=str)
    data = {}
    if check_exist(batch_id,course_id,subject_id,attendance_date):
        data = {'statuscode':2,'message': 'Attendance already exist do you want to update?'}
    else:
        data = {'statuscode':1,'message': 'No attendance found, do you want to take attendance?'}
    return jsonify(data)


def check_exist(batch_id,course_id,subject_id,attendance_date):
    flag = False
    attendance = get_db().execute(
        'SELECT id'
        ' FROM attendance '
        ' WHERE batch_id = ? AND course_id = ? AND subject_id = ? AND attendance_date = ?',
        (batch_id,course_id,subject_id,attendance_date)
    ).fetchone()

    if attendance is not None:
        flag = True
    return flag


@bp.route('/getstudents', methods=['GET'])
def getstudents():
    batch_id = request.args.get('batch_id', default='default_value', type=str)
    course_id = request.args.get('course_id', default='default_value', type=str)
    subject_id = request.args.get('subject_id', default='default_value', type=str)
    attendance_date = request.args.get('attendance_date', default='default_value', type=str)
    students = []
    if check_exist(batch_id,course_id,subject_id,attendance_date):
        students = get_db().execute(
                '''SELECT DISTINCT s.id, s.first_name, s.last_name, IFNULL(ast.status, "") AS status
                    FROM student s
                    JOIN batch b ON b.id = s.batch_id 
                    JOIN course c ON c.id = s.course_id 
                    JOIN course_subject cs ON cs.course_id = c.id 
                    JOIN subject sb ON sb.id = cs.subject_id 
                    JOIN teacher t ON t.id = sb.teacher_id
                    LEFT JOIN attendance a ON a.batch_id = b.id AND a.course_id = c.id AND a.subject_id = sb.id 
                    LEFT JOIN attendance_student ast ON ast.attendance_id = a.id AND s.id = ast.student_id
                    WHERE s.batch_id = ?  AND cs.course_id = ? AND cs.subject_id = ? AND t.user_id = ? AND a.attendance_date = ? ORDER BY s.id''',
                (batch_id,course_id,subject_id,g.user['id'],attendance_date)
            ).fetchall()
    else:
        students = get_db().execute(
                '''SELECT DISTINCT s.id, s.first_name, s.last_name, "0" AS status
                    FROM student s
                    JOIN batch b ON b.id = s.batch_id 
                    JOIN course c ON c.id = s.course_id 
                    JOIN course_subject cs ON cs.course_id = c.id 
                    JOIN subject sb ON sb.id = cs.subject_id 
                    JOIN teacher t ON t.id = sb.teacher_id
                    WHERE s.batch_id = ?  AND cs.course_id = ? AND cs.subject_id = ? AND t.user_id = ? ORDER BY s.id''',
                (batch_id,course_id,subject_id,g.user['id'])
            ).fetchall()

    student_ids = [{'id': student['id'],'first_name': student['first_name'],'last_name': student['last_name'],'status': student['status']} for student in students]
    return jsonify(student_ids)


@bp.route('/submitattendance', methods=['POST'])
def submitattendance():
    response = {'statuscode':0,'message':'error in saving data','data':None}
    if request.method == 'POST':
        data = request.json
        batch_id = data['batch_id']
        course_id = data['course_id']
        subject_id = data['subject_id']
        attendance_date = data['attendance_date']
        studentdata = data['studentdata']
        if check_exist(batch_id,course_id,subject_id,attendance_date):
            attendance_id = get_db().execute(
                            'SELECT id'
                            ' FROM attendance '
                            ' WHERE batch_id = ? AND course_id = ? AND subject_id = ? AND attendance_date = ?',
                            (batch_id,course_id,subject_id,attendance_date)
                        ).fetchone()
            db = get_db()
            cursor = db.cursor()
            for student in studentdata:
                cursor.execute(
                '''
                UPDATE attendance_student
                SET status = ?
                WHERE attendance_id = ? AND student_id = ?;
                ''',
                (student['status'],attendance_id['id'], student['id'])
                )
                db.commit()
            flash("successfully updated student attendance","success")
            response = {'statuscode':2,'message': 'Attendance updated successfully','data':attendance_id['id']}
        else:
            db = get_db()
            cursor = db.cursor()
            
            cursor.execute(
                '''
                INSERT INTO attendance (batch_id, course_id, subject_id, attendance_date, user_id, type)
                VALUES (?, ?, ?, ?, ?, ?)
                ''',
                (batch_id,course_id,subject_id, attendance_date,g.user['id'],'manual')
            )
            db.commit()
            inserted_id = cursor.lastrowid
            for student in studentdata:
                cursor.execute(
                'Insert into attendance_student(attendance_id,student_id,status)'
                ' VALUES (?, ?, ?);',
                (inserted_id, student['id'], student['status'])
                )
                db.commit()
            flash("successfully saved student attendance","success")
            response = {'statuscode':1,'message':'successfully saved student attendance','data':inserted_id}
    return jsonify(response)


@bp.route('/<int:id>/download-excel', methods=['GET'])
def download_excel(id):
    excel_file = get_create_excel_sheet(id)
    return send_file(excel_file, as_attachment=True)


def get_create_excel_sheet(id):
    file_path = ''
    if id > 0 :
        relative_file_path = 'attendance_sheets'
        
        attendance_detail = get_attendance_detail(id)
        attendance_date = attendance_detail['attendance_date']
        batch_year = attendance_detail['batch_year']
        course_name = attendance_detail['course_name']
        subject_name = attendance_detail['subject_name']

        relative_file_path = relative_file_path+'\\'+batch_year+'\\'+course_name+'\\'+subject_name+'\\'+attendance_date+'.xlsx'

        file_path = os.path.join(PROJECT_ROOT, relative_file_path)

        db = get_db()
        cursor = db.cursor()
        data = cursor.execute(
            ''' select st.id "Roll no.",
                st.first_name || " "|| st.last_name "Student Name",
                CASE 
                    WHEN ast.status = 1 THEN 'Present'
                    ELSE 'Absent' END
                Status
                from attendance a 
                JOIN attendance_student ast ON a.id = ast.attendance_id
                JOIN batch b ON b.id = a.batch_id
                JOIN course c ON c.id = a.course_id
                JOIN subject s ON s.id = a.subject_id
                JOIN student st ON c.id = st.course_id AND ast.student_id = st.id
                JOIN teacher t ON t.user_id = a.user_id 
                WHERE a.id = ?''',
                (id,)
        ).fetchall()
        wb = Workbook()
        ws = wb.active
        # Write column headers to the first row of the worksheet
        headers = [description[0] for description in cursor.description]
        for col_index, header in enumerate(headers, start=1):
            ws.cell(row=1, column=col_index, value=header)

        # Write data to the worksheet
        for row_index, row in enumerate(data, start=2):
            for col_index, value in enumerate(row, start=1):
                ws.cell(row=row_index, column=col_index, value=value)

        wb.save(file_path)
       
    return file_path


def get_attendance_detail(id):
    
    attendance = get_db().execute(
        '''
        SELECT a.attendance_date,b.batch_year,c.course_name,s.subject_name 
        FROM attendance a 
        JOIN batch b ON b.id = a.batch_id
        JOIN course c ON c.id = a.course_id
        JOIN subject  s ON s.id = a.subject_id
        WHERE a.id =?
        ''',
        (id,)
    ).fetchone()

    return attendance


@bp.route('/send-mail', methods=['POST'])
def send_attendance_mail():
    response = {'statuscode':0,'message':'error in saving data','data':None}
    if request.method == 'POST':
        data = request.json
        attendance_id = data['attendance_id']
        teacher_email = data['mail']
        sender_email = EMAIL
        sender_password = PASSWORD
        receiver_email = teacher_email
        subject = 'Student' + 's Attendance '
        body = 'Please find attachment for attendance of student'
        attachment_path = get_create_excel_sheet(int(attendance_id))
        print(attachment_path)
        file_name= attachment_path.split('\\')[-1]
        print(file_name)
        send_email(sender_email, sender_password, receiver_email, subject, body, attachment_path,file_name)
        
        response = {'statuscode':1,'message':'successfully send mail to mail id :'+ teacher_email,'data':teacher_email}
    return jsonify(response)


def send_email(sender_email, sender_password, receiver_email, subject, body, attachment_path,file_name):
    message = MIMEMultipart()
    message['From'] = sender_email
    message['To'] = receiver_email
    message['Subject'] = subject

    message.attach(MIMEText(body, 'plain'))

    with open(attachment_path, 'rb') as attachment:
        part = MIMEBase('application', 'octet-stream')
        part.set_payload(attachment.read())
    
    encoders.encode_base64(part)
    part.add_header('Content-Disposition', f'attachment; filename= {file_name}')

    message.attach(part)
    text = message.as_string()

    server = smtplib.SMTP('smtp.gmail.com', 587)
    server.starttls()
    server.login(sender_email, sender_password)

    server.sendmail(sender_email, receiver_email, text)

    server.quit()
